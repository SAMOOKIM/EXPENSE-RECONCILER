import os, re, sys, threading, webbrowser, tempfile
from pathlib import Path
from flask import Flask, render_template, request, jsonify, send_file
import openpyxl
import fitz
from PIL import Image
import pytesseract
from itertools import combinations
from difflib import SequenceMatcher

APP = Flask(__name__)
BASE = Path(getattr(sys, "_MEIPASS", Path(__file__).resolve().parent))
BUNDLED_TESS = BASE / "tesseract" / "tesseract.exe"
BUNDLED_TESSDATA = BASE / "tesseract" / "tessdata"

if BUNDLED_TESS.exists():
    pytesseract.pytesseract.tesseract_cmd = str(BUNDLED_TESS)
    os.environ["TESSDATA_PREFIX"] = str(BUNDLED_TESSDATA)

REPORT_DIR = Path(tempfile.gettempdir()) / "expense_reconciler_reports"
REPORT_DIR.mkdir(parents=True, exist_ok=True)

def norm_text(s):
    if s is None:
        return ""
    s = str(s).lower()
    s = re.sub(r"[\s\u3000\-\_\(\)\[\]{}'\"`·.,/\\:;]+", "", s)
    return s

def merchant_score(a, b):
    a, b = norm_text(a), norm_text(b)
    if not a or not b:
        return 0.0
    if a in b or b in a:
        return 1.0
    return SequenceMatcher(None, a, b).ratio()

def parse_amounts(text):
    # Prefer amounts with comma separators or currency-like context.
    vals = []
    for m in re.finditer(r'(?<!\d)(\d{1,3}(?:,\d{3})+|\d{4,8})(?!\d)', text):
        raw = m.group(1).replace(",", "")
        try:
            n = int(raw)
        except ValueError:
            continue
        if 100 <= n <= 100_000_000:
            vals.append(n)
    return vals

def parse_dates(text):
    out = []
    patterns = [
        r'(20\d{2})[.\-/](\d{1,2})[.\-/](\d{1,2})',
        r'(20\d{2})(\d{2})(\d{2})',
    ]
    for pat in patterns:
        for m in re.finditer(pat, text):
            out.append(f"{int(m.group(1)):04d}-{int(m.group(2)):02d}-{int(m.group(3)):02d}")
    return list(dict.fromkeys(out))

def read_excel(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    idx = {str(v).strip(): i for i, v in enumerate(headers) if v is not None}
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        vals = list(r)
        amount = vals[idx.get("거래금액", idx.get("장부금액", 0))] or 0
        try:
            amount = int(round(float(amount)))
        except Exception:
            amount = 0
        desc = vals[idx.get("적요", 0)] if idx.get("적요", 0) < len(vals) else ""
        merchant = vals[idx.get("거래처명", idx.get("거래처", 0))] if idx.get("거래처명", idx.get("거래처", 0)) < len(vals) else ""
        no = vals[idx.get("No", 0)] if idx.get("No", 0) < len(vals) else len(rows)+1
        dates = parse_dates(str(desc))
        rows.append({
            "no": no, "amount": amount, "merchant": str(merchant or ""),
            "desc": str(desc or ""), "date": dates[0] if dates else ""
        })
    # This workbook has summary rows followed by card-detail rows.
    # For reconciliation, use rows that look like actual evidence rows when available.
    detail = [x for x in rows if x["amount"] and ("신용카드" in x["desc"] or x["merchant"])]
    return detail or rows

def pdf_extract(path):
    doc = fitz.open(path)
    pages = []
    ocr_used = False
    for i, page in enumerate(doc, 1):
        text = page.get_text("text") or ""
        if len(re.sub(r"\s+", "", text)) < 30:
            ocr_used = True
            try:
                pix = page.get_pixmap(matrix=fitz.Matrix(2.0, 2.0), alpha=False)
                img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
                text = pytesseract.image_to_string(img, lang="kor+eng", config="--psm 6", timeout=45)
            except Exception as e:
                text = f"[OCR_ERROR] {e}"
        pages.append({"page": i, "text": text, "amounts": parse_amounts(text), "dates": parse_dates(text)})
    return pages, ocr_used

def best_amount_for_row(row, pages):
    target = row["amount"]
    candidates = []
    for p in pages:
        date_score = 1.0 if (not row["date"] or row["date"] in p["dates"]) else 0.0
        for a in p["amounts"]:
            diff = abs(a-target)
            score = (1.0 if a == target else 0.0) + 0.35*date_score + max(0, 1-diff/max(target,1))*0.25
            candidates.append((score, p["page"], a))
    candidates.sort(reverse=True)
    return candidates[:10]

def subset_match(target, values, max_items=6):
    vals = [(p, a) for p, a in values if 0 < a <= target]
    # Avoid combinatorial explosion.
    vals = vals[:24]
    best = None
    for k in range(2, min(max_items, len(vals))+1):
        for combo in combinations(vals, k):
            s = sum(a for _, a in combo)
            if s == target:
                return list(combo)
            if best is None or abs(target-s) < best[0]:
                best = (abs(target-s), list(combo))
    return None

def reconcile(rows, pages):
    results = []
    for row in rows:
        exact = []
        all_values = []
        for p in pages:
            for a in p["amounts"]:
                all_values.append((p["page"], a))
                if a == row["amount"]:
                    exact.append((p["page"], a))
        # Prefer exact amount on a page whose date matches.
        if exact:
            dated = [(p,a) for p,a in exact if not row["date"] or row["date"] in pages[p-1]["dates"]]
            chosen = dated or exact
            results.append({**row, "status":"MATCH", "matched_pages":sorted(set(p for p,_ in chosen)),
                            "matched_amounts":[a for _,a in chosen[:6]], "note":"동일 금액 증빙 확인"})
            continue

        combo = subset_match(row["amount"], all_values, max_items=6)
        if combo:
            results.append({**row, "status":"SUM_MATCH",
                            "matched_pages":sorted(set(p for p,_ in combo)),
                            "matched_amounts":[a for _,a in combo],
                            "note":"여러 PDF 금액의 합계가 Excel 금액과 일치"})
        else:
            near = best_amount_for_row(row, pages)
            near_txt = ", ".join(f"p{p}:{a:,}" for _,p,a in near[:3])
            results.append({**row, "status":"MISSING", "matched_pages":[],
                            "matched_amounts":[], "note":f"직접/합산 일치 없음. 후보: {near_txt}"})
    return results

def write_report(results, out):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "대조결과"
    headers = ["No","상태","거래일","거래처","Excel 금액","PDF 페이지","PDF 금액","비고","적요"]
    ws.append(headers)
    for r in results:
        ws.append([r["no"], r["status"], r["date"], r["merchant"], r["amount"],
                   ", ".join(map(str,r.get("matched_pages",[]))),
                   " + ".join(map(str,r.get("matched_amounts",[]))),
                   r.get("note",""), r.get("desc","")])
    for c in ws.columns:
        ws.column_dimensions[c[0].column_letter].width = min(max(len(str(x.value or "")) for x in c)+2, 60)
    wb.save(out)

@APP.get("/")
def index():
    return render_template("index.html")

@APP.post("/api/reconcile")
def api_reconcile():
    try:
        ex = request.files["excel"]
        pdf = request.files["pdf"]
        with tempfile.TemporaryDirectory() as td:
            ep = Path(td)/ex.filename
            pp = Path(td)/pdf.filename
            ex.save(ep); pdf.save(pp)
            rows = read_excel(ep)
            pages, ocr = pdf_extract(pp)
            results = reconcile(rows, pages)
            report = REPORT_DIR / "reconciliation_result.xlsx"
            write_report(results, report)
        summary = {
            "excel_rows": len(rows), "pdf_pages": len(pages),
            "match": sum(r["status"]=="MATCH" for r in results),
            "sum_match": sum(r["status"]=="SUM_MATCH" for r in results),
            "missing": sum(r["status"]=="MISSING" for r in results),
            "excel_total": sum(r["amount"] for r in rows),
            "ocr_used": ocr
        }
        return jsonify({"summary":summary, "results":results, "report_url":"/download"})
    except Exception as e:
        return jsonify({"error":str(e)}), 500

@APP.get("/download")
def download():
    return send_file(REPORT_DIR/"reconciliation_result.xlsx", as_attachment=True,
                     download_name="증빙대조결과.xlsx")

def open_browser():
    webbrowser.open("http://127.0.0.1:5179/")

if __name__ == "__main__":
    threading.Timer(1.0, open_browser).start()
    APP.run(host="127.0.0.1", port=5179, debug=False)
